from openpyxl import load_workbook

# Load the workbook
workbook = load_workbook("C:\\Users\\91735\\Desktop\\demo.xlsx")

# Extract data from Sheet1
sheet1 = workbook['Sheet1']
data_sheet1 = []

for row in sheet1.iter_rows():
    data_row = []
    for cell in row:
        data_row.append(cell.value)
    data_sheet1.append(data_row)

# Extract data from Sheet2
sheet2 = workbook['Sheet2']
data_sheet2 = []

for row in sheet2.iter_rows():
    data_row = []
    for cell in row:
      data_row.append(cell.value)
    data_sheet2.append(data_row)

# Printing the data from Sheet1
print("Data from Sheet1:")
for row in data_sheet1:
    print(row)

# Printing the data from Sheet2
print("Data from Sheet2:")
for row in data_sheet2:
    print(row)





































# note
# here one excel sheet is having 2 sheets named 'sheet1' and 'sheet2'
# we are accessing the worksheet using "openpyxl" this is the python excel package
# so after that we are looping over the 'sheet1' and 'sheet2'